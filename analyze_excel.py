"""
Script para analizar el template de Excel y encontrar las celdas correctas.
Guarda la salida en un archivo.
"""

from openpyxl import load_workbook

TEMPLATE_PATH = r"C:\Users\raulc\Downloads\debug biostar para checadores\modulo mobper\F-RH-18-MIT-FORMATO-DE-MOVIMIENTO-DE-PERSONAL-3(1).xlsx"
OUTPUT_FILE = r"C:\Users\raulc\Downloads\debug biostar para checadores\excel_analysis.txt"

def analyze_template():
    output_lines = []
    
    def log(msg):
        print(msg)
        output_lines.append(msg)
    
    log("="*60)
    log("ANALISIS DEL TEMPLATE MOBPER")
    log("="*60)
    
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws = wb.active
    
    log(f"\nHoja activa: {ws.title}")
    log(f"Dimensiones: {ws.dimensions}")
    log(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Listar celdas fusionadas
    log("\n" + "-"*60)
    log("CELDAS FUSIONADAS (MERGED):")
    log("-"*60)
    for merged in ws.merged_cells.ranges:
        log(f"  {merged}")
    
    # Mostrar contenido
    log("\n" + "-"*60)
    log("CONTENIDO DE CELDAS (primeras 35 filas):")
    log("-"*60)
    
    for row in range(1, 36):
        row_data = []
        for col in range(1, 15):  # A hasta N
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val and str(val).strip():
                col_letter = chr(64 + col)  # A=1, B=2, etc
                val_str = str(val)[:40]  # Limitar longitud
                row_data.append(f"{col_letter}{row}='{val_str}'")
        
        if row_data:
            log(f"Fila {row}: {', '.join(row_data)}")
    
    wb.close()
    log("\n" + "="*60)
    
    # Guardar en archivo
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_lines))
    
    print(f"\nResultados guardados en: {OUTPUT_FILE}")

if __name__ == '__main__':
    analyze_template()
