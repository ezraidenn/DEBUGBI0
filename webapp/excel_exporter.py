"""
Exportador de pase de lista a Excel con diseño profesional.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pytz
import io


class EmergencyExcelExporter:
    """Exportador profesional de pase de lista a Excel."""
    
    # Colores corporativos - Paleta café elegante
    COLOR_HEADER = "3E2723"  # Café oscuro
    COLOR_SUBHEADER = "5D4037"  # Café medio
    COLOR_PRESENT = "4CAF50"  # Verde natural
    COLOR_ABSENT = "8D6E63"  # Café medio (ausentes)
    COLOR_PENDING = "D7CCC8"  # Beige claro
    COLOR_ALT_ROW = "F5F5F5"  # Gris muy claro para filas alternas
    
    def __init__(self, emergency, roll_call_data):
        """
        Inicializar exportador.
        
        Args:
            emergency: Objeto EmergencySession
            roll_call_data: Dict con grupos y miembros del pase de lista
        """
        self.emergency = emergency
        self.roll_call_data = roll_call_data
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Pase de Lista"
        
    def _create_border(self, style='thin'):
        """Crear borde para celdas."""
        side = Side(style=style, color="000000")
        return Border(left=side, right=side, top=side, bottom=side)
    
    def _write_header(self):
        """Escribir encabezado del documento."""
        # Título principal
        self.ws.merge_cells('A1:G1')
        cell = self.ws['A1']
        cell.value = "PASE DE LISTA DE EMERGENCIA"
        cell.font = Font(name='Arial', size=18, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[1].height = 30
        
        # Información de la emergencia
        row = 3
        # Convertir fechas a zona horaria de México
        mexico_tz = pytz.timezone('America/Mexico_City')
        started_at_local = self.emergency.started_at.replace(tzinfo=pytz.UTC).astimezone(mexico_tz)
        
        info_data = [
            ("Zona:", self.emergency.zone.name),
            ("Tipo:", self.emergency.emergency_type.upper()),
            ("Iniciada por:", self.emergency.started_by_user.full_name or self.emergency.started_by_user.username),
            ("Fecha de inicio:", started_at_local.strftime('%d/%m/%Y %H:%M:%S')),
            ("Estado:", "RESUELTA" if self.emergency.status == 'resolved' else "ACTIVA"),
        ]
        
        if self.emergency.resolved_at:
            resolved_at_local = self.emergency.resolved_at.replace(tzinfo=pytz.UTC).astimezone(mexico_tz)
            info_data.append(("Fecha de resolución:", resolved_at_local.strftime('%d/%m/%Y %H:%M:%S')))
        
        for label, value in info_data:
            self.ws[f'A{row}'] = label
            self.ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
            self.ws[f'B{row}'] = value
            self.ws[f'B{row}'].font = Font(name='Arial', size=11)
            self.ws.merge_cells(f'B{row}:G{row}')
            row += 1
        
        # Estadísticas
        row += 1
        stats = self.roll_call_data.get('stats', {})
        self.ws.merge_cells(f'A{row}:G{row}')
        cell = self.ws[f'A{row}']
        cell.value = "ESTADÍSTICAS"
        cell.font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=self.COLOR_SUBHEADER, end_color=self.COLOR_SUBHEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        self.ws.row_dimensions[row].height = 25
        
        row += 1
        stats_row = row
        self.ws[f'A{stats_row}'] = "Total"
        self.ws[f'B{stats_row}'] = stats.get('total', 0)
        self.ws[f'C{stats_row}'] = "Presentes"
        self.ws[f'D{stats_row}'] = stats.get('present', 0)
        self.ws[f'E{stats_row}'] = "Ausentes"
        self.ws[f'F{stats_row}'] = stats.get('absent', 0)
        self.ws[f'G{stats_row}'] = f"Pendientes: {stats.get('pending', 0)}"
        
        for col in ['A', 'C', 'E', 'G']:
            self.ws[f'{col}{stats_row}'].font = Font(name='Arial', size=10, bold=True)
        for col in ['B', 'D', 'F']:
            self.ws[f'{col}{stats_row}'].font = Font(name='Arial', size=10)
            self.ws[f'{col}{stats_row}'].alignment = Alignment(horizontal='center')
        
        return row + 2
    
    def _write_table_header(self, start_row):
        """Escribir encabezado de la tabla."""
        headers = ["Grupo", "Nombre", "ID BioStar", "Estado", "Marcado por", "Fecha/Hora", "Notas"]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_SUBHEADER, end_color=self.COLOR_SUBHEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._create_border()
        
        self.ws.row_dimensions[start_row].height = 20
        return start_row + 1
    
    def _write_data_rows(self, start_row):
        """Escribir filas de datos."""
        current_row = start_row
        
        for group in self.roll_call_data.get('groups', []):
            group_name = group['group_name']
            group_color = group.get('group_color', '#6c757d')
            
            for idx, member in enumerate(group['members']):
                # Alternar color de fondo
                if idx % 2 == 0:
                    bg_fill = PatternFill(start_color=self.COLOR_ALT_ROW, end_color=self.COLOR_ALT_ROW, fill_type="solid")
                else:
                    bg_fill = PatternFill(fill_type=None)
                
                # Grupo
                cell = self.ws.cell(row=current_row, column=1, value=group_name)
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.fill = bg_fill
                cell.border = self._create_border()
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Nombre
                cell = self.ws.cell(row=current_row, column=2, value=member['user_name'])
                cell.font = Font(name='Arial', size=10)
                cell.fill = bg_fill
                cell.border = self._create_border()
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # ID BioStar
                cell = self.ws.cell(row=current_row, column=3, value=member['biostar_user_id'])
                cell.font = Font(name='Arial', size=10)
                cell.fill = bg_fill
                cell.border = self._create_border()
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Estado
                status = member['status']
                status_text = {
                    'present': 'PRESENTE',
                    'absent': 'AUSENTE',
                    'pending': 'PENDIENTE'
                }.get(status, status.upper())
                
                status_color = {
                    'present': self.COLOR_PRESENT,
                    'absent': self.COLOR_ABSENT,
                    'pending': '6D4C41'  # Café claro para pendientes
                }.get(status, '95A5A6')
                
                cell = self.ws.cell(row=current_row, column=4, value=status_text)
                cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                cell.border = self._create_border()
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Marcado por
                marked_by = member.get('marked_by', 'N/A')
                cell = self.ws.cell(row=current_row, column=5, value=marked_by)
                cell.font = Font(name='Arial', size=10)
                cell.fill = bg_fill
                cell.border = self._create_border()
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Fecha/Hora (convertir a zona horaria México UTC-6)
                if member.get('marked_at'):
                    try:
                        # Parsear timestamp UTC
                        marked_dt = datetime.fromisoformat(member['marked_at'].replace('Z', '+00:00'))
                        # Convertir a zona horaria de México (UTC-6)
                        mexico_tz = pytz.timezone('America/Mexico_City')
                        marked_dt_local = marked_dt.astimezone(mexico_tz)
                        marked_str = marked_dt_local.strftime('%d/%m/%Y %H:%M')
                    except:
                        marked_str = member['marked_at']
                else:
                    marked_str = 'Sin marcar'
                
                cell = self.ws.cell(row=current_row, column=6, value=marked_str)
                cell.font = Font(name='Arial', size=10)
                cell.fill = bg_fill
                cell.border = self._create_border()
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Notas
                notes = member.get('notes', '')
                cell = self.ws.cell(row=current_row, column=7, value=notes)
                cell.font = Font(name='Arial', size=9, italic=True)
                cell.fill = bg_fill
                cell.border = self._create_border()
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                current_row += 1
        
        return current_row
    
    def _adjust_column_widths(self):
        """Ajustar anchos de columna."""
        column_widths = {
            'A': 20,  # Grupo
            'B': 35,  # Nombre
            'C': 15,  # ID BioStar
            'D': 15,  # Estado
            'E': 20,  # Marcado por
            'F': 18,  # Fecha/Hora
            'G': 30,  # Notas
        }
        
        for col, width in column_widths.items():
            self.ws.column_dimensions[col].width = width
    
    def _add_footer(self, last_row):
        """Agregar pie de página."""
        footer_row = last_row + 2
        self.ws.merge_cells(f'A{footer_row}:G{footer_row}')
        cell = self.ws[f'A{footer_row}']
        # Usar zona horaria de México
        mexico_tz = pytz.timezone('America/Mexico_City')
        now_local = datetime.now(mexico_tz)
        cell.value = f"Documento generado el {now_local.strftime('%d/%m/%Y a las %H:%M:%S')}"
        cell.font = Font(name='Arial', size=9, italic=True, color="7F8C8D")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def generate(self):
        """Generar el archivo Excel completo."""
        # Escribir secciones
        data_start_row = self._write_header()
        table_start_row = self._write_table_header(data_start_row)
        last_data_row = self._write_data_rows(table_start_row)
        
        # Ajustes finales
        self._adjust_column_widths()
        self._add_footer(last_data_row)
        
        # Retornar archivo en memoria
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_emergency(emergency, roll_call_data):
        """
        Método estático para exportar una emergencia.
        
        Args:
            emergency: Objeto EmergencySession
            roll_call_data: Dict con grupos y miembros
            
        Returns:
            BytesIO con el archivo Excel
        """
        exporter = EmergencyExcelExporter(emergency, roll_call_data)
        return exporter.generate()
