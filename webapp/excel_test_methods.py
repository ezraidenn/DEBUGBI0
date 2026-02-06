"""
Módulo de prueba CORREGIDO para edición de Excel.
Usa las celdas correctas basadas en el análisis del template.

CELDAS IDENTIFICADAS:
- E8:M8 = Nombre del empleado (escribir en E8)
- H10:K10 = Fecha de autorización
- P10:R10 = Fecha de aplicación  
- G20:R21 = Motivo (escribir en G20)
- F27:J27 = Departamento
- K27:M27 = A Departamento
- E29:R30 = Motivo modificación
- C32:D33 = Puesto Actual
- K32:L33 = Puesto Nuevo
"""

import os
import shutil
from datetime import datetime

# Ruta del template original
TEMPLATE_PATH = r"C:\Users\raulc\Downloads\debug biostar para checadores\modulo mobper\F-RH-18-MIT-FORMATO-DE-MOVIMIENTO-DE-PERSONAL-3(1).xlsx"

# Directorio de salida
OUTPUT_DIR = r"C:\Users\raulc\Downloads\debug biostar para checadores\webapp\output\excel_tests"

# Datos de prueba
TEST_DATA = {
    'nombre': 'JUAN PEREZ GARCIA - TEST',
    'departamento': 'SISTEMAS DE INFORMACION',
    'puesto': 'Desarrollador Senior',
    'numero_empleado': '12345',
    'fecha_autorizacion': '06/02/2026',
    'fecha_aplicacion': '01 al 15 de Febrero de 2026',
    'motivo': 'Retardo los dias 03, 05, 08 de Febrero 2026. PRUEBA DE ESCRITURA.',
}

def ensure_output_dir():
    """Crear directorio de salida si no existe."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    return OUTPUT_DIR

def get_output_path(method_name):
    """Obtener ruta de salida para un método específico."""
    ensure_output_dir()
    timestamp = datetime.now().strftime('%H%M%S')
    return os.path.join(OUTPUT_DIR, f"{method_name}_{timestamp}.xlsx")


# ============================================================
# MÉTODO 1: openpyxl - Unmerge/Merge approach
# ============================================================
def method_openpyxl_basic():
    """
    openpyxl: Des-fusionar, escribir, re-fusionar.
    """
    method_name = "openpyxl_basic"
    result = {
        'method': method_name,
        'library': 'openpyxl (unmerge)',
        'description': 'Des-fusiona celdas, escribe, y re-fusiona.',
        'success': False,
        'output_file': None,
        'error': None
    }
    
    try:
        from openpyxl import load_workbook
        
        output_path = get_output_path(method_name)
        shutil.copy(TEMPLATE_PATH, output_path)
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Guardar rangos fusionados de las celdas a editar
        ranges_to_edit = {
            'E8:M8': ('E8', TEST_DATA['nombre']),
            'G20:R21': ('G20', TEST_DATA['motivo']),
        }
        
        for merge_range, (cell, value) in ranges_to_edit.items():
            # Des-fusionar
            if merge_range in [str(m) for m in ws.merged_cells.ranges]:
                ws.unmerge_cells(merge_range)
            # Escribir
            ws[cell] = value
            # Re-fusionar
            ws.merge_cells(merge_range)
        
        wb.save(output_path)
        wb.close()
        
        result['success'] = True
        result['output_file'] = output_path
        
    except Exception as e:
        result['error'] = str(e)
    
    return result


# ============================================================
# MÉTODO 2: openpyxl - Escribir en celda principal directamente
# ============================================================
def method_openpyxl_preserve():
    """
    openpyxl: Escribir directamente en la celda principal del rango.
    """
    method_name = "openpyxl_preserve"
    result = {
        'method': method_name,
        'library': 'openpyxl (direct)',
        'description': 'Escribe solo en la celda principal (top-left) del rango fusionado.',
        'success': False,
        'output_file': None,
        'error': None
    }
    
    try:
        from openpyxl import load_workbook
        
        output_path = get_output_path(method_name)
        shutil.copy(TEMPLATE_PATH, output_path)
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Para celdas fusionadas, debemos escribir en la celda superior-izquierda
        # La celda E8 es la principal de E8:M8
        # La celda G20 es la principal de G20:R21
        
        # Primero, encontrar la celda real (no MergedCell)
        for merged_range in ws.merged_cells.ranges:
            if 'E8' in str(merged_range):
                # Obtener la celda principal
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                ws.cell(row=min_row, column=min_col).value = TEST_DATA['nombre']
            
            if 'G20' in str(merged_range):
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                ws.cell(row=min_row, column=min_col).value = TEST_DATA['motivo']
        
        wb.save(output_path)
        wb.close()
        
        result['success'] = True
        result['output_file'] = output_path
        
    except Exception as e:
        result['error'] = str(e)
    
    return result


# ============================================================
# MÉTODO 3: openpyxl full con rich_text
# ============================================================
def method_openpyxl_full():
    """
    openpyxl: Con rich_text=True.
    """
    method_name = "openpyxl_full"
    result = {
        'method': method_name,
        'library': 'openpyxl full',
        'description': 'openpyxl con rich_text=True y acceso directo.',
        'success': False,
        'output_file': None,
        'error': None
    }
    
    try:
        from openpyxl import load_workbook
        
        output_path = get_output_path(method_name)
        shutil.copy(TEMPLATE_PATH, output_path)
        
        wb = load_workbook(output_path, rich_text=True)
        ws = wb.active
        
        # Escribir en las celdas principales
        ws.cell(row=8, column=5).value = TEST_DATA['nombre']  # E8
        ws.cell(row=20, column=7).value = TEST_DATA['motivo']  # G20
        
        wb.save(output_path)
        wb.close()
        
        result['success'] = True
        result['output_file'] = output_path
        
    except Exception as e:
        result['error'] = str(e)
    
    return result


# ============================================================
# MÉTODO 4: Spire.XLS con Range
# ============================================================
def method_spire_xls():
    """
    Spire.XLS: Usando Range["E8"].Text con celdas correctas.
    """
    method_name = "spire_xls"
    result = {
        'method': method_name,
        'library': 'Spire.XLS Range',
        'description': 'Spire.XLS usando Range["E8"].Text con celdas correctas.',
        'success': False,
        'output_file': None,
        'error': None
    }
    
    try:
        from spire.xls import Workbook
        
        output_path = get_output_path(method_name)
        
        workbook = Workbook()
        workbook.LoadFromFile(TEMPLATE_PATH)
        sheet = workbook.Worksheets[0]
        
        # Escribir en las celdas correctas
        sheet.Range["E8"].Text = TEST_DATA['nombre']
        sheet.Range["G20"].Text = TEST_DATA['motivo']
        sheet.Range["F27"].Text = TEST_DATA['departamento']
        
        workbook.SaveToFile(output_path)
        workbook.Dispose()
        
        result['success'] = True
        result['output_file'] = output_path
        
    except Exception as e:
        result['error'] = str(e)
    
    return result


# ============================================================
# MÉTODO 5: Spire.XLS con índices
# ============================================================
def method_spire_getcell():
    """
    Spire.XLS: Usando índices de fila/columna.
    """
    method_name = "spire_getcell"
    result = {
        'method': method_name,
        'library': 'Spire.XLS Index',
        'description': 'Spire.XLS usando sheet[row, col] con celdas correctas.',
        'success': False,
        'output_file': None,
        'error': None
    }
    
    try:
        from spire.xls import Workbook
        
        output_path = get_output_path(method_name)
        
        workbook = Workbook()
        workbook.LoadFromFile(TEMPLATE_PATH)
        sheet = workbook.Worksheets[0]
        
        # E=5, G=7, F=6 (1-indexed)
        # Usar indexación directa
        sheet[8, 5].Text = TEST_DATA['nombre']   # E8
        sheet[20, 7].Text = TEST_DATA['motivo']  # G20
        sheet[27, 6].Text = TEST_DATA['departamento']  # F27
        
        workbook.SaveToFile(output_path)
        workbook.Dispose()
        
        result['success'] = True
        result['output_file'] = output_path
        
    except Exception as e:
        result['error'] = str(e)
    
    return result


# ============================================================
# MÉTODO 6: Spire.XLS con Value
# ============================================================
def method_spire_value():
    """
    Spire.XLS: Usando .Value en vez de .Text con celdas correctas.
    """
    method_name = "spire_value"
    result = {
        'method': method_name,
        'library': 'Spire.XLS Value',
        'description': 'Spire.XLS usando .Value con celdas correctas.',
        'success': False,
        'output_file': None,
        'error': None
    }
    
    try:
        from spire.xls import Workbook
        
        output_path = get_output_path(method_name)
        
        workbook = Workbook()
        workbook.LoadFromFile(TEMPLATE_PATH)
        sheet = workbook.Worksheets[0]
        
        # Usar Value en vez de Text
        sheet.Range["E8"].Value = TEST_DATA['nombre']
        sheet.Range["G20"].Value = TEST_DATA['motivo']
        sheet.Range["F27"].Value = TEST_DATA['departamento']
        
        workbook.SaveToFile(output_path)
        workbook.Dispose()
        
        result['success'] = True
        result['output_file'] = output_path
        
    except Exception as e:
        result['error'] = str(e)
    
    return result


# ============================================================
# MÉTODO 7: xlwings
# ============================================================
def method_xlwings():
    """
    xlwings: Controla Excel directamente con celdas correctas.
    """
    method_name = "xlwings"
    result = {
        'method': method_name,
        'library': 'xlwings',
        'description': 'xlwings controlando Excel via COM con celdas correctas.',
        'success': False,
        'output_file': None,
        'error': None
    }
    
    try:
        import xlwings as xw
        
        output_path = get_output_path(method_name)
        shutil.copy(TEMPLATE_PATH, output_path)
        
        app = xw.App(visible=False)
        wb = app.books.open(output_path)
        ws = wb.sheets[0]
        
        # Escribir en celdas correctas
        ws.range('E8').value = TEST_DATA['nombre']
        ws.range('G20').value = TEST_DATA['motivo']
        ws.range('F27').value = TEST_DATA['departamento']
        
        wb.save()
        wb.close()
        app.quit()
        
        result['success'] = True
        result['output_file'] = output_path
        
    except Exception as e:
        result['error'] = str(e)
    
    return result


# ============================================================
# MÉTODO 8: win32com
# ============================================================
def method_win32com():
    """
    win32com: Control directo de Excel via COM con celdas correctas.
    """
    method_name = "win32com"
    result = {
        'method': method_name,
        'library': 'win32com',
        'description': 'PyWin32 controlando Excel via COM con celdas correctas.',
        'success': False,
        'output_file': None,
        'error': None
    }
    
    try:
        import win32com.client as win32
        
        output_path = get_output_path(method_name)
        shutil.copy(TEMPLATE_PATH, output_path)
        
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        
        wb = excel.Workbooks.Open(output_path)
        ws = wb.ActiveSheet
        
        # Escribir en celdas correctas
        ws.Range('E8').Value = TEST_DATA['nombre']
        ws.Range('G20').Value = TEST_DATA['motivo']
        ws.Range('F27').Value = TEST_DATA['departamento']
        
        wb.Save()
        wb.Close()
        excel.Quit()
        
        result['success'] = True
        result['output_file'] = output_path
        
    except Exception as e:
        result['error'] = str(e)
    
    return result


# ============================================================
# MÉTODO 9: pandas + openpyxl
# ============================================================
def method_pandas():
    """
    pandas + openpyxl: Con celdas correctas.
    """
    method_name = "pandas_openpyxl"
    result = {
        'method': method_name,
        'library': 'pandas openpyxl',
        'description': 'Pandas usando openpyxl engine con celdas correctas.',
        'success': False,
        'output_file': None,
        'error': None
    }
    
    try:
        from openpyxl import load_workbook
        
        output_path = get_output_path(method_name)
        shutil.copy(TEMPLATE_PATH, output_path)
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Escribir usando row/col
        ws.cell(row=8, column=5).value = TEST_DATA['nombre']
        ws.cell(row=20, column=7).value = TEST_DATA['motivo']
        ws.cell(row=27, column=6).value = TEST_DATA['departamento']
        
        wb.save(output_path)
        wb.close()
        
        result['success'] = True
        result['output_file'] = output_path
        
    except Exception as e:
        result['error'] = str(e)
    
    return result


# ============================================================
# MÉTODO 10: xlutils (no funciona)
# ============================================================
def method_xlutils():
    """
    xlrd/xlwt/xlutils: Solo para .xls legacy.
    """
    method_name = "xlutils"
    result = {
        'method': method_name,
        'library': 'xlutils',
        'description': 'Solo soporta .xls, no .xlsx.',
        'success': False,
        'output_file': None,
        'error': "Esta libreria solo soporta archivos .xls, no .xlsx"
    }
    
    return result


# ============================================================
# Ejecutar todas las pruebas
# ============================================================
def run_all_tests():
    """Ejecutar todos los métodos de prueba."""
    methods = [
        method_openpyxl_basic,
        method_openpyxl_preserve,
        method_openpyxl_full,
        method_spire_xls,
        method_spire_getcell,
        method_spire_value,
        method_xlwings,
        method_win32com,
        method_pandas,
        method_xlutils,
    ]
    
    results = []
    for method in methods:
        print(f"Ejecutando: {method.__name__}...")
        result = method()
        results.append(result)
        status = "✓" if result['success'] else "✗"
        print(f"  {status} {result['method']}: {result.get('error', 'OK')}")
    
    return results


if __name__ == '__main__':
    results = run_all_tests()
    print("\n" + "="*60)
    print("RESUMEN:")
    print("="*60)
    for r in results:
        status = "OK" if r['success'] else "FAIL"
        print(f"[{status}] {r['library']}: {r.get('output_file', r.get('error', 'N/A'))}")
