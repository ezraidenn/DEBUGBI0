"""
Modulo de generacion de formato Excel para MovPer
Usa win32com (pywin32) para llenar el template Excel preservando completamente el formato y shapes.
"""

import win32com.client as win32
from datetime import datetime, date
import os
import shutil
import tempfile
from datetime import datetime, timedelta
import pythoncom
from typing import List, Dict, Tuple

# Directorio base
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Configuracion de rutas
TEMPLATE_PATH = os.path.join(BASE_DIR, 'templates', 'F-RH-18-MIT-FORMATO-DE-MOVIMIENTO-DE-PERSONAL-3(1).xlsx')

# =============================================================================
# MAPEO DE CELDAS DEL TEMPLATE EXCEL
# =============================================================================
# Basado en analisis visual del formato F-RH-18-MIT

CELDAS = {
    # Encabezado - Coordenadas exactas confirmadas por el usuario
    'NOMBRE': 'E8',              # Nombre del empleado (fila 8, columnas E-M)
    'DEPARTAMENTO': 'Q8',        # Departamento (fila 8, columnas Q-R)
    'FECHA_AUTORIZACION': 'H10', # Fecha de autorización (fila 10, columnas H-K)
    'FECHA_APLICACION': 'P10',   # Fechas de aplicación (fila 10, columnas P-R)
    
    # Círculos de tipo de permiso
    'PARA_FALTAR': 'D15',        # Checkbox para faltar (fila 15, columna D)
    'PARA_SALIR_REGRESAR': 'D16', # Checkbox para salir y regresar (fila 16, columna D)
    'PARA_LLEGAR_TARDE': 'D17',  # Checkbox para llegar tarde (fila 17, columna D)
    'PARA_RETIRARSE': 'M15',     # Checkbox retirarse temprano (fila 15, columna M)
    'OLVIDO_CHECAR': 'M16',      # Checkbox olvidó checar (fila 16, columna M)
    
    # Goce de sueldo (necesitaría confirmar columnas exactas)
    'GOCE_SI': 'F17',            # Círculo SI
    'GOCE_NO': 'G17',            # Círculo NO
    
    # Motivo
    'MOTIVO': 'G20',             # Campo de motivo (filas 20-21, columnas G-R)
    
    # Firmas (SOLICITUD DE PERMISO)
    'SOLICITO_NOMBRE': 'E56',    # Nombre del solicitante (fila 56, columnas E-G)
    'AUTORIZO_NOMBRE': 'J56',    # Nombre del autorizador/jefe (fila 56, columna J)
    'RECIBIO_NOMBRE': 'Q56',     # "Recursos Humanos" (fila 56, estimado)

    # AVISO DE VACACIONES (filas 39-52)
    'VAC_FECHA_INGRESO':      'E41',   # Fecha de ingreso (no disponible)
    'VAC_ANTIGUEDAD':         'L41',   # Antigüedad (no disponible)
    'VAC_PERIODO':            'Q41',   # Período vacacional
    'VAC_DIAS_CORRESPONDEN':  'F43',   # Días que corresponden
    'VAC_DIAS_EFECTIVOS':     'K43',   # Días efectivos de vacaciones
    'VAC_DIAS_PENDIENTES':    'P43',   # Días pendientes (merge top-left P43)
    'VAC_FECHA_SALIDA':       'G45',   # Fecha de inicio de vacaciones (merge G45:I45)
    'VAC_FECHA_REGRESO':      'P45',   # Fecha de regreso
    'VAC_NOMBRE_SUPLENTE':    'F47',   # Nombre del suplente
    'VAC_DEPARTAMENTO':       'P47',   # Departamento
    'VAC_NOTA':               'F49',   # Nota
    'VAC_SOLICITO':           'E56',   # Firma solicitó (misma fila que permiso)
    'VAC_AUTORIZO':           'J56',   # Firma autorizó
    'VAC_RECIBIO':            'N56',   # Firma recibió (RH)
}

# =============================================================================
# MAPEO DE SHAPES (CÍRCULOS) DEL TEMPLATE
# =============================================================================
# Basado en analisis con win32com - indices de shapes

SHAPES = {
    'PARA_FALTAR': 1,           # Shape #1 en C15
    'PARA_SALIR_REGRESAR': 2,   # Shape #2 en C16
    'OLVIDO_CHECAR': 3,         # Shape #3 en M16
    'PARA_RETIRARSE': 4,        # Shape #4 en L14
    'PARA_LLEGAR_TARDE': 5,     # Shape #5 en C17
    'GOCE_SI': 9,               # Shape #9 en P17
    'GOCE_NO': 10,              # Shape #10 en R17
}

# Colores RGB para Excel (formato BGR)
def rgb_to_excel(r, g, b):
    return r + (g << 8) + (b << 16)

COLOR_NEGRO = rgb_to_excel(0, 0, 0)
COLOR_BLANCO = rgb_to_excel(255, 255, 255)

# Meses en español
MESES_ES = {
    1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
    5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
    9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
}


# =============================================================================
# FUNCIONES AUXILIARES
# =============================================================================

# =============================================================================
# ADAPTACION INTELIGENTE DE FUENTE - basada en mediciones reales del template
# =============================================================================
#
# Dimensiones medidas con win32com (en puntos):
#
#  CAMPO               CELDA       ANCHO_PT  FONT_BASE  FUENTE         WRAP
#  NOMBRE              E8:M8         27.00      12      Arial Narrow   No
#  DEPARTAMENTO        Q8:R8         17.25      12      Arial Narrow   No
#  FECHA_AUTORIZACION  H10:K10       26.25      10      Arial Narrow   No
#  FECHA_APLICACION    P10:R10       32.25       8      Arial Narrow   No
#  MOTIVO              G20:R21       37.50       9      Arial          Si (2 filas)
#  SOLICITO_NOMBRE     E56:G56       27.00      10      Arial Narrow   No
#  AUTORIZO_NOMBRE     J56           35.25      10      Arial Narrow   No
#  RECIBIO_NOMBRE      N56:R56       17.25      10      Arial Narrow   No
#
# Calibracion de caracteres por punto:
#   Arial Narrow es ~30% mas estrecha que Arial regular.
#   A tamaño 10pt, Arial Narrow: ~1.35 chars/pt de ancho de celda
#   A tamaño 10pt, Arial regular: ~1.05 chars/pt de ancho de celda
#   La capacidad escala linealmente: chars_max = (ancho_pt * chars_por_pt * 10) / font_size
#
# Ejemplo: NOMBRE (27pt, Arial Narrow):
#   font 12 -> 27 * 1.35 * 10/12 = ~30 chars
#   font 10 -> 27 * 1.35 * 10/10 = ~36 chars
#   font  8 -> 27 * 1.35 * 10/8  = ~45 chars

# Configuracion de cada campo:
#   (col_width_total, font_max, font_min, wrap_lines)
#
# col_width_total = suma de ColumnWidth de cada columna del merge area
#   (ColumnWidth = chars de '0' en fuente Normal del doc = Calibri 11)
#
# Mediciones reales del template (win32com ColumnWidth sumado):
#   NOMBRE            E8:M8    = 40.66  chars
#   DEPARTAMENTO      Q8:R8    = 11.86  chars
#   FECHA_AUTH        H10:K10  = 21.71  chars
#   FECHA_APLICACION  P10:R10  = 17.29  chars
#   MOTIVO            G20:R21  = ~65.0  chars (12 cols, wrap=2 filas)
#   SOLICITO_NOMBRE   E56:G56  = 11.53  chars
#   AUTORIZO_NOMBRE   J56      =  6.00  chars
#   RECIBIO_NOMBRE    N56:R56  = 27.43  chars
#
# Formula de capacidad:
#   Arial Narrow es ~78% del ancho de Calibri 11 (fuente base de ColumnWidth)
#   chars_reales(font_size) = col_width / RATIO_NARROW * (11.0 / font_size) * wrap_lines
#   donde RATIO_NARROW = 0.78
#
# Ejemplo DEPARTAMENTO (11.86 chars):
#   font 12 -> 11.86 / 0.78 * (11/12) = ~13.9 chars
#   font  9 -> 11.86 / 0.78 * (11/9)  = ~18.6 chars
#   font  7 -> 11.86 / 0.78 * (11/7)  = ~23.9 chars

RATIO_NARROW = 0.88   # Arial Narrow vs Calibri 11 (conservador para merged cells)
RATIO_ARIAL  = 1.10   # Arial regular vs Calibri 11 (conservador para merged cells)

CAMPO_CONFIG = {
    #                    col_width  fmax  fmin  wrap  ratio
    'NOMBRE':             (40.66,   10,  10,    1,   RATIO_NARROW),
    'DEPARTAMENTO':       (11.86,   14,   6,    1,   RATIO_NARROW),
    'FECHA_AUTORIZACION': (21.71,   10,  10,    1,   RATIO_NARROW),
    'FECHA_APLICACION':   (17.29,   10,   6,    1,   RATIO_NARROW),
    'MOTIVO':             (65.00,   11,   7,    2,   RATIO_ARIAL),
    'SOLICITO_NOMBRE':    (11.53,   12,   6,    1,   RATIO_NARROW),
    'AUTORIZO_NOMBRE':    ( 6.00,   10,   6,    1,   RATIO_NARROW),
    'RECIBIO_NOMBRE':     (27.43,   12,   7,    1,   RATIO_NARROW),
}


def calcular_font_adaptativo(texto: str, campo: str) -> int:
    """
    Calcula el tamaño de fuente optimo para que el texto quepa en el campo.

    Usa ColumnWidth total del merge area (medido empiricamente) como base,
    con la relacion de ancho entre la fuente del campo y Calibri 11.

    Formula: chars_reales = col_width / ratio * (11.0 / font_size) * wrap_lines

    Args:
        texto: Texto a escribir.
        campo: Clave en CAMPO_CONFIG.

    Returns:
        int: Tamaño de fuente en puntos.
    """
    if not texto:
        cfg = CAMPO_CONFIG.get(campo)
        return cfg[1] if cfg else 10

    config = CAMPO_CONFIG.get(campo)
    if not config:
        length = len(texto)
        if length <= 15:
            return 10
        elif length <= 22:
            return 9
        else:
            return 8

    col_width, font_max, font_min, wrap_lines, ratio = config
    n_chars = len(texto)

    for font_size in range(font_max, font_min - 1, -1):
        capacidad = col_width / ratio * (11.0 / font_size) * wrap_lines
        if n_chars <= capacidad:
            return font_size

    return font_min


def calcular_tamano_letra(texto: str, max_size: int = 12, min_size: int = 8) -> int:
    """
    Wrapper de compatibilidad. Usa calcular_font_adaptativo cuando se conoce
    el campo; de lo contrario aplica logica generica por longitud.
    """
    length = len(texto)
    for font_size in range(max_size, min_size - 1, -1):
        capacidad = 30 * (11.0 / font_size)
        if length <= capacidad:
            return font_size
    return min_size


# Campos que NO deben usar ShrinkToFit porque el texto desborda visualmente
# hacia celdas adyacentes vacías (comportamiento natural de Excel).
NO_SHRINK_FIELDS = {'AUTORIZO_NOMBRE', 'SOLICITO_NOMBRE'}


def set_cell_text(sheet, celda: str, texto: str, campo: str):
    """
    Escribe texto en una celda aplicando tamaño de fuente adaptativo.
    Si el texto no cabe ni a font_min, activa ShrinkToFit como respaldo,
    excepto para campos en NO_SHRINK_FIELDS que desbordan naturalmente.

    Args:
        sheet: Worksheet de Excel (win32com)
        celda: Direccion de celda, ej. 'E8'
        texto: Texto a escribir
        campo: Clave en CAMPO_CONFIG para calcular el font size
    """
    rng = sheet.Range(celda)
    font_size = calcular_font_adaptativo(texto, campo)
    rng.Value = texto
    rng.Font.Size = font_size

    if campo in NO_SHRINK_FIELDS:
        rng.ShrinkToFit = False
    else:
        config = CAMPO_CONFIG.get(campo)
        if config:
            col_width, font_max, font_min, wrap_lines, ratio = config
            cap_min = col_width / ratio * (11.0 / font_min) * wrap_lines
            if len(texto) > cap_min:
                rng.ShrinkToFit = True
                print(f"[MOVPER EXCEL] {campo} '{texto}' ({len(texto)}c) -> ShrinkToFit activado (cap_min={cap_min:.1f})")
            else:
                rng.ShrinkToFit = False
    print(f"[MOVPER EXCEL] {campo} '{texto}' ({len(texto)} chars) -> font {font_size}pt")


def formatear_fechas_compactas(dias: List[int], mes: str) -> str:
    """
    Formatea los dias de forma compacta.
    Si hay más de 4 días, usa formato sin espacios: 1,2,3,5,8
    Si hay 4 o menos, usa espacios: 1, 2, 3, 5
    """
    if len(dias) <= 4:
        dias_str = ', '.join(str(d) for d in sorted(dias))
    else:
        dias_str = ','.join(str(d) for d in sorted(dias))
    
    return f"{dias_str} de {mes}"


def set_circle_color(sheet, shape_index: int, is_selected: bool):
    """
    Establece el color de un círculo/shape.
    Negro = seleccionado, Blanco = no seleccionado.
    """
    try:
        shape = sheet.Shapes(shape_index)
        if is_selected:
            shape.Fill.ForeColor.RGB = COLOR_NEGRO
            shape.Fill.Visible = True
        else:
            shape.Fill.ForeColor.RGB = COLOR_BLANCO
            shape.Fill.Visible = False  # Solo contorno
        print(f"[MOVPER EXCEL] Shape {shape_index}: {'NEGRO' if is_selected else 'BLANCO'}")
    except Exception as e:
        print(f"[MOVPER EXCEL] Error en shape {shape_index}: {e}")


def reemplazar_logo(sheet, logo_path: str) -> bool:
    """
    Reemplaza el logo en el Excel (Shape 25 - Picture 12).
    Preserva la posición original y mantiene el aspect ratio del nuevo logo.
    
    Args:
        sheet: Worksheet de Excel
        logo_path: Ruta absoluta al archivo de logo
    
    Returns:
        bool: True si se reemplazó exitosamente
    """
    try:
        logo_abs_path = os.path.abspath(logo_path)
        
        if not os.path.exists(logo_abs_path):
            print(f"[MOVPER EXCEL] Logo no encontrado: {logo_abs_path}")
            return False
        
        print(f"[MOVPER EXCEL] Reemplazando logo con: {logo_abs_path}")
        
        # Obtener propiedades del logo original MIT (Shape 25)
        original_logo = sheet.Shapes(25)
        logo_left = original_logo.Left
        logo_top = original_logo.Top
        mit_width = original_logo.Width   # 83.4 puntos
        mit_height = original_logo.Height  # 45.1 puntos
        
        # Límites de expansión
        max_height = mit_height           # Límite vertical = altura MIT
        max_width = mit_width * 1.5       # Límite horizontal = ancho MIT × 1.5
        
        print(f"[MOVPER EXCEL] Límites: Height={max_height:.1f}, Width={max_width:.1f} (MIT × 1.5)")
        
        # Eliminar logo existente
        original_logo.Delete()
        print(f"[MOVPER EXCEL] Logo anterior eliminado (Shape 25)")
        
        # Insertar nuevo logo en la misma posición temporal
        picture = sheet.Shapes.AddPicture(
            Filename=logo_abs_path,
            LinkToFile=False,
            SaveWithDocument=True,
            Left=logo_left,
            Top=logo_top,
            Width=-1,  # -1 = usar dimensiones originales del archivo
            Height=-1
        )
        
        # Bloquear aspect ratio para que se mantenga la proporción original del logo
        picture.LockAspectRatio = -1  # True en COM
        
        # Obtener dimensiones actuales del logo (originales del archivo)
        original_width = picture.Width
        original_height = picture.Height
        
        # Calcular factores de escala para cada límite
        scale_by_height = max_height / original_height
        scale_by_width = max_width / original_width
        
        # Usar el factor MENOR para que el logo quepa sin exceder ningún límite
        # (expandir hasta tocar el primer límite)
        scale_factor = min(scale_by_height, scale_by_width)
        
        print(f"[MOVPER EXCEL] Factores: height={scale_by_height:.2f}, width={scale_by_width:.2f}, usando={scale_factor:.2f}")
        
        # Aplicar escala (con aspect ratio bloqueado, ambas dimensiones se ajustan proporcionalmente)
        picture.Width = original_width * scale_factor
        
        # Centrar verticalmente dentro del espacio disponible
        final_height = picture.Height
        vertical_space = max_height
        vertical_offset = (vertical_space - final_height) / 2
        picture.Top = logo_top + vertical_offset
        
        print(f"[MOVPER EXCEL] Centrado vertical: offset={vertical_offset:.1f}, Top={picture.Top:.1f}")
        
        print(f"[MOVPER EXCEL] Escalado aplicado: original=({original_width:.1f}x{original_height:.1f}) -> final=({picture.Width:.1f}x{picture.Height:.1f})")
        
        # Quitar bordes y sombras (como el original)
        picture.Line.Visible = 0  # Sin borde
        try:
            picture.Shadow.Visible = 0  # Sin sombra
        except:
            pass  # Algunos logos pueden no tener esta propiedad
        
        print(f"[MOVPER EXCEL] Logo reemplazado: Width={picture.Width}, Height={picture.Height}")
        return True
        
    except Exception as e:
        print(f"[MOVPER EXCEL] Error reemplazando logo: {e}")
        import traceback
        import sys
        traceback.print_exc(file=sys.stderr)
        return False


def agrupar_dias_consecutivos(dias: List[int]) -> str:
    """
    Agrupa dias consecutivos para mostrar de forma compacta.
    Ejemplo: [1,2,3,5,7,8,9] -> "1-3, 5, 7-9"
    """
    if not dias:
        return ""
    
    dias = sorted(set(dias))
    grupos = []
    inicio = dias[0]
    fin = dias[0]
    
    for dia in dias[1:]:
        if dia == fin + 1:
            fin = dia
        else:
            if inicio == fin:
                grupos.append(str(inicio))
            else:
                grupos.append(f"{inicio}-{fin}" if fin - inicio > 1 else f"{inicio},{fin}")
            inicio = fin = dia
    
    # Agregar ultimo grupo
    if inicio == fin:
        grupos.append(str(inicio))
    else:
        grupos.append(f"{inicio}-{fin}" if fin - inicio > 1 else f"{inicio},{fin}")
    
    return ", ".join(grupos)


def filtrar_incidencias_a_justificar(incidencias: List[Dict]) -> List[Dict]:
    """
    Filtra las incidencias que deben aparecer en el formato MovPer.
    
    Incluye:
    - Retardos que tienen justificado=True (o no definido, por defecto True)
    - Faltas que tienen clasificacion (REMOTO, GUARDIA, PERMISO)
      EXCEPTO INCAPACIDAD y VACACIONES (vacaciones van en su propio AVISO)
    - Salidas tempranas justificadas (salida_estado == 'SALIDA_TEMPRANA' y salida_justificado)
    - Olvido de checada (entrada_no_checada o salida_no_checada) con olvido_checar_justificado=True
    
    Excluye:
    - A_TIEMPO, INHABIL, DESCANSO
    - Retardos con justificado=False
    - Faltas sin clasificar
    - Faltas clasificadas como INCAPACIDAD o VACACIONES
    - Salidas tempranas con salida_justificado=False
    """
    justificadas = []
    fechas_incluidas = set()
    
    for inc in incidencias:
        estado_auto = inc.get('estado_auto', '')
        clasificacion = inc.get('clasificacion', '')
        justificado = inc.get('justificado', True)
        salida_estado = inc.get('salida_estado', '')
        salida_justificado = inc.get('salida_justificado', True)
        entrada_no_checada = inc.get('entrada_no_checada', False)
        salida_no_checada = inc.get('salida_no_checada', False)
        fecha = inc.get('fecha')
        incluir = False
        
        # Retardos: solo los justificados
        if estado_auto == 'RETARDO' and justificado:
            incluir = True
        
        # Faltas: solo las clasificadas (excepto INCAPACIDAD, VACACIONES, ERROR_SISTEMA)
        # VACACIONES van en su propio AVISO DE VACACIONES separado
        # ERROR_SISTEMA = no es falta real, excluir del formato
        elif estado_auto == 'FALTA' and clasificacion:
            if clasificacion not in ('INCAPACIDAD', 'VACACIONES', 'ERROR_SISTEMA'):
                incluir = True
        
        # Salidas tempranas justificadas
        if salida_estado == 'SALIDA_TEMPRANA' and salida_justificado:
            incluir = True
        
        # Olvido de checada: solo si justificado (False = error del sistema, excluir)
        olvido_justificado = inc.get('olvido_checar_justificado', True)
        if (entrada_no_checada or salida_no_checada) and olvido_justificado:
            incluir = True
        
        if incluir and fecha not in fechas_incluidas:
            justificadas.append(inc)
            fechas_incluidas.add(fecha)
    
    return justificadas


def agrupar_periodos_vacaciones(incidencias: List[Dict]) -> List[Dict]:
    """
    Agrupa dias de vacaciones en periodos consecutivos.
    Dias no laborales (DESCANSO, INHABIL) entre vacaciones se incluyen en el periodo.

    Returns:
        Lista de periodos: [{'dias': [date,...], 'fecha_salida': date, 'fecha_regreso': date}]
    """
    from datetime import timedelta

    # Extraer solo los dias clasificados como VACACIONES
    dias_vac = sorted([
        inc['fecha'] for inc in incidencias
        if inc.get('clasificacion') == 'VACACIONES' and inc.get('estado_auto') == 'FALTA'
    ])

    if not dias_vac:
        return []

    periodos = []
    grupo = [dias_vac[0]]

    for i in range(1, len(dias_vac)):
        prev = dias_vac[i - 1]
        curr = dias_vac[i]
        # Si la diferencia es <= 7 dias (permite fin de semana + feriado entre periodos)
        if (curr - prev).days <= 7:
            grupo.append(curr)
        else:
            periodos.append(grupo)
            grupo = [curr]
    periodos.append(grupo)

    result = []
    for grupo in periodos:
        fecha_salida = min(grupo)
        # fecha_regreso = dia siguiente al ultimo dia de vacaciones
        fecha_regreso = max(grupo) + timedelta(days=1)
        result.append({
            'dias': grupo,
            'fecha_salida': fecha_salida,
            'fecha_regreso': fecha_regreso,
            'dias_efectivos': len(grupo),
        })
    return result


def generar_aviso_vacaciones(
    user,
    preset,
    periodo_vac: Dict,
    quincena: Dict,
) -> Tuple[str, str]:
    """
    Genera un AVISO DE VACACIONES llenando la seccion inferior del template Excel.

    Args:
        user: MobPerUser
        preset: PresetUsuario
        periodo_vac: Dict con 'dias', 'fecha_salida', 'fecha_regreso', 'dias_efectivos'
        quincena: Dict con info de la quincena

    Returns:
        (output_path, filename)
    """
    import shutil, tempfile

    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template no encontrado: {TEMPLATE_PATH}")

    # Crear copia temporal
    tmp_dir = tempfile.mkdtemp()
    nombre = preset.nombre_formato if preset and preset.nombre_formato else user.nombre_completo
    nombre_display = nombre.title()
    safe_name = nombre_display.replace(' ', '_')[:30]
    fecha_salida = periodo_vac['fecha_salida']
    filename = f"VACACIONES_{safe_name}_{fecha_salida.strftime('%Y%m%d')}.xlsx"
    output_path = os.path.join(tmp_dir, filename)
    shutil.copy2(TEMPLATE_PATH, output_path)

    pythoncom.CoInitialize()
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(os.path.abspath(output_path))
        sheet = wb.ActiveSheet

        # --- Logo ---
        if preset and preset.company and preset.company.logo_filename:
            logo_path = os.path.join(BASE_DIR, 'static', 'logos', preset.company.logo_filename)
            if os.path.exists(logo_path):
                reemplazar_logo(sheet, logo_path)

        # --- Encabezado superior (mismo que MovPer) ---
        set_cell_text(sheet, CELDAS['NOMBRE'], nombre_display, 'NOMBRE')
        depto = preset.departamento_formato if preset and preset.departamento_formato else ''
        set_cell_text(sheet, CELDAS['DEPARTAMENTO'], depto.upper(), 'DEPARTAMENTO')
        now = datetime.now()
        mes_corto = {1:'ene',2:'feb',3:'mar',4:'abr',5:'may',6:'jun',7:'jul',8:'ago',
                     9:'sep',10:'oct',11:'nov',12:'dic'}
        fecha_auth = f"{now.day:02d}-{mes_corto[now.month]}-{now.strftime('%y')}"
        set_cell_text(sheet, CELDAS['FECHA_AUTORIZACION'], fecha_auth, 'FECHA_AUTORIZACION')

        # --- Limpiar seccion SOLICITUD DE PERMISO (viene con datos del template base) ---
        sheet.Range(CELDAS['MOTIVO']).Value = ''
        # Limpiar todos los shapes (circulos) - dejarlos sin relleno
        for shape_idx in SHAPES.values():
            try:
                set_circle_color(sheet, shape_idx, False)
            except Exception:
                pass
        # Limpiar FECHA_APLICACION del encabezado (se llenara con fechas de vacaciones)
        sheet.Range(CELDAS['FECHA_APLICACION']).Value = ''

        # --- FECHA_APLICACION: dias de vacaciones ---
        dias_ef = periodo_vac['dias_efectivos']
        f_salida = periodo_vac['fecha_salida']
        # fecha_regreso = ultimo dia de vacaciones (NO el dia siguiente)
        f_regreso = max(periodo_vac['dias'])

        # Fechas de aplicacion: lista de dias de vacaciones
        dias_vac_nums = sorted([d.day for d in periodo_vac['dias']])
        mes_nombre = mes_corto[f_salida.month]
        if len(dias_vac_nums) == 1:
            fecha_aplic_str = f"{dias_vac_nums[0]:02d}-{mes_nombre}-{str(f_salida.year)[2:]}"
        else:
            dias_str = ','.join(str(d) for d in dias_vac_nums)
            fecha_aplic_str = f"{dias_str} {mes_nombre}-{str(f_salida.year)[2:]}"
        set_cell_text(sheet, CELDAS['FECHA_APLICACION'], fecha_aplic_str, 'FECHA_APLICACION')

        # --- AVISO DE VACACIONES ---
        # Dias efectivos
        sheet.Range(CELDAS['VAC_DIAS_EFECTIVOS']).Value = dias_ef
        sheet.Range(CELDAS['VAC_DIAS_EFECTIVOS']).Font.Size = 10

        # Fecha de salida: dd-mmm-yy (primer dia de vacaciones)
        f_sal_str = f"{f_salida.day:02d}-{mes_corto[f_salida.month]}-{str(f_salida.year)[2:]}"
        # Escribir directamente en la celda del merge area
        rng_sal = sheet.Range(CELDAS['VAC_FECHA_SALIDA'])
        rng_sal.Value = f_sal_str
        rng_sal.Font.Size = 10
        print(f"[MOVPER VACACIONES] VAC_FECHA_SALIDA={CELDAS['VAC_FECHA_SALIDA']} -> '{f_sal_str}'")

        # Fecha de regreso: dd-mmm-yy (ultimo dia de vacaciones)
        f_reg_str = f"{f_regreso.day:02d}-{mes_corto[f_regreso.month]}-{str(f_regreso.year)[2:]}"
        rng_reg = sheet.Range(CELDAS['VAC_FECHA_REGRESO'])
        rng_reg.Value = f_reg_str
        rng_reg.Font.Size = 10
        print(f"[MOVPER VACACIONES] VAC_FECHA_REGRESO={CELDAS['VAC_FECHA_REGRESO']} -> '{f_reg_str}'")

        # Departamento
        set_cell_text(sheet, CELDAS['VAC_DEPARTAMENTO'], depto.upper(), 'DEPARTAMENTO')

        # Firmas
        set_cell_text(sheet, CELDAS['VAC_SOLICITO'], nombre_display, 'SOLICITO_NOMBRE')
        jefe = preset.jefe_directo_nombre if preset and preset.jefe_directo_nombre else 'PENDIENTE'
        set_cell_text(sheet, CELDAS['VAC_AUTORIZO'], jefe.upper(), 'AUTORIZO_NOMBRE')
        set_cell_text(sheet, CELDAS['VAC_RECIBIO'], 'RECURSOS HUMANOS', 'RECIBIO_NOMBRE')

        print(f"[MOVPER VACACIONES] Periodo {f_salida} - {f_regreso}: {dias_ef} dias")

        wb.Save()
        wb.Close(False)
        print(f"[MOVPER VACACIONES] Guardado: {output_path}")
        return output_path, filename

    except Exception as e:
        print(f"[MOVPER VACACIONES] Error: {e}")
        import traceback
        traceback.print_exc()
        try:
            wb.Close(False)
        except Exception:
            pass
        raise
    finally:
        try:
            excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


def generar_formato_excel(
    user,
    preset,
    incidencias: List[Dict],
    quincena: Dict,
    con_goce: bool = True
) -> Tuple[str, str]:
    """
    Genera el formato de movimiento de personal editando la plantilla Excel.
    
    Args:
        user: MovPerUser - Usuario actual
        preset: PresetUsuario - Configuracion del usuario
        incidencias: List[Dict] - Lista de incidencias clasificadas
        quincena: Dict - Info de la quincena {inicio, fin, nombre}
        con_goce: bool - Si es con goce de sueldo
    
    Returns:
        Tuple[str, str]: (ruta_archivo, nombre_archivo)
    """
    
    # Crear archivo temporal para el Excel (se borra automáticamente después de usarse)
    fecha_str = quincena['inicio'].strftime('%Y%m%d')
    output_filename = f"MovPer_{user.numero_socio}_{fecha_str}.xlsx"
    
    # Crear archivo temporal con nombre descriptivo
    temp_file = tempfile.NamedTemporaryFile(
        mode='wb',
        suffix='.xlsx',
        prefix=f'MovPer_{user.numero_socio}_',
        delete=False  # No borrar automáticamente, lo haremos manualmente después de enviarlo
    )
    output_path = temp_file.name
    temp_file.close()
    
    print(f"[MOVPER EXCEL] Iniciando generación de Excel con win32com")
    print(f"[MOVPER EXCEL] Template: {TEMPLATE_PATH}")
    print(f"[MOVPER EXCEL] Output temporal: {output_path}")
    print(f"[MOVPER EXCEL] Usuario: {user.nombre_completo}")
    print(f"[MOVPER EXCEL] Incidencias: {len(incidencias)}")
    
    # IMPORTANTE: Copiar template PRIMERO, luego abrir la copia
    print(f"[MOVPER EXCEL] Copiando template a archivo temporal...")
    shutil.copy(TEMPLATE_PATH, output_path)
    
    # Inicializar COM
    pythoncom.CoInitialize()
    
    # Usar gencache.EnsureDispatch (mas estable que Dispatch)
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False  # NO mostrar Excel
    excel.DisplayAlerts = False
    
    try:
        # Abrir la COPIA (no el template original)
        wb = excel.Workbooks.Open(os.path.abspath(output_path))
        sheet = wb.ActiveSheet
        print(f"[MOVPER EXCEL] Archivo abierto con win32com")
        
        # =============================================================
        # REEMPLAZAR LOGO SEGÚN EMPRESA DEL USUARIO
        # =============================================================
        if preset and preset.company and preset.company.logo_filename:
            logo_path = os.path.join(BASE_DIR, 'static', 'logos', preset.company.logo_filename)
            if os.path.exists(logo_path):
                print(f"[MOVPER EXCEL] Empresa: {preset.company.name}")
                reemplazar_logo(sheet, logo_path)
            else:
                print(f"[MOVPER EXCEL] Logo no encontrado para empresa {preset.company.name}: {logo_path}")
        else:
            print(f"[MOVPER EXCEL] Sin empresa configurada, usando logo por defecto (MIT)")
        
        # =============================================================
        # FILTRAR INCIDENCIAS QUE VAN A JUSTIFICARSE
        # =============================================================
        # Solo incluir: retardos justificados + faltas clasificadas (excepto INCAPACIDAD)
        incidencias_justificadas = filtrar_incidencias_a_justificar(incidencias)
        print(f"[MOVPER EXCEL] Incidencias totales: {len(incidencias)}")
        print(f"[MOVPER EXCEL] Incidencias a justificar: {len(incidencias_justificadas)}")
        
        # =============================================================
        # LLENAR DATOS DEL ENCABEZADO
        # =============================================================
        
        # Nombre del empleado (Title Case: Raúl Abel Cetina Pool)
        nombre = preset.nombre_formato if preset and preset.nombre_formato else user.nombre_completo
        nombre_display = nombre.title()
        set_cell_text(sheet, CELDAS['NOMBRE'], nombre_display, 'NOMBRE')
        
        # Departamento
        departamento = preset.departamento_formato if preset and preset.departamento_formato else "Sin departamento"
        depto_upper = departamento.upper()
        set_cell_text(sheet, CELDAS['DEPARTAMENTO'], depto_upper, 'DEPARTAMENTO')
        
        # Fecha de autorizacion (hoy) en formato dd-mmm-yy con mes en espanol
        now = datetime.now()
        mes_corto = {1:'ene',2:'feb',3:'mar',4:'abr',5:'may',6:'jun',7:'jul',8:'ago',9:'sep',10:'oct',11:'nov',12:'dic'}
        fecha_auth = f"{now.day:02d}-{mes_corto[now.month]}-{now.strftime('%y')}"
        set_cell_text(sheet, CELDAS['FECHA_AUTORIZACION'], fecha_auth, 'FECHA_AUTORIZACION')
        
        # Fecha de aplicacion (dias del periodo)
        dias_periodo = construir_fechas_aplicacion(incidencias_justificadas, quincena)
        set_cell_text(sheet, CELDAS['FECHA_APLICACION'], dias_periodo, 'FECHA_APLICACION')
        
        # =============================================================
        # ANALIZAR TIPOS DE INCIDENCIAS Y MARCAR CIRCULOS REALES
        # =============================================================
        
        tipos = analizar_tipos_incidencias(incidencias_justificadas)
        
        # Primero, poner todos los circulos en blanco (no seleccionado)
        print(f"[MOVPER EXCEL] Reseteando todos los circulos a blanco...")
        for shape_name, shape_idx in SHAPES.items():
            set_circle_color(sheet, shape_idx, False)
        
        # Marcar los circulos correspondientes segun tipos encontrados
        print(f"[MOVPER EXCEL] Marcando círculos según incidencias...")
        # PARA FALTAR: cualquier tipo de falta (genérica, remoto, guardia, permiso, vacaciones)
        tiene_faltas = tipos['faltas'] or tipos['remotos'] or tipos['guardias'] or tipos['permisos'] or tipos['vacaciones']
        if tiene_faltas:
            set_circle_color(sheet, SHAPES['PARA_FALTAR'], True)
        if tipos['retardos']:
            set_circle_color(sheet, SHAPES['PARA_LLEGAR_TARDE'], True)
        if tipos['salir_regresar']:
            set_circle_color(sheet, SHAPES['PARA_SALIR_REGRESAR'], True)
        if tipos['olvido_checar']:
            set_circle_color(sheet, SHAPES['OLVIDO_CHECAR'], True)
        if tipos['retirarse_temprano']:
            set_circle_color(sheet, SHAPES['PARA_RETIRARSE'], True)
        
        # Goce de sueldo
        if con_goce:
            set_circle_color(sheet, SHAPES['GOCE_SI'], True)
            set_circle_color(sheet, SHAPES['GOCE_NO'], False)
        else:
            set_circle_color(sheet, SHAPES['GOCE_SI'], False)
            set_circle_color(sheet, SHAPES['GOCE_NO'], True)
        
        # =============================================================
        # GENERAR TEXTO DEL MOTIVO
        # =============================================================
        
        motivo = generar_texto_motivo(incidencias_justificadas, tipos)
        set_cell_text(sheet, CELDAS['MOTIVO'], motivo, 'MOTIVO')
        
        # =============================================================
        # FIRMAS
        # =============================================================
        
        # Solicito (el empleado) - mismo nombre que encabezado
        set_cell_text(sheet, CELDAS['SOLICITO_NOMBRE'], nombre_display, 'SOLICITO_NOMBRE')
        
        # Autorizo (jefe directo)
        jefe = preset.jefe_directo_nombre if preset and preset.jefe_directo_nombre else "PENDIENTE"
        jefe_upper = jefe.upper()
        set_cell_text(sheet, CELDAS['AUTORIZO_NOMBRE'], jefe_upper, 'AUTORIZO_NOMBRE')
        
        # Recibio (RH)
        set_cell_text(sheet, CELDAS['RECIBIO_NOMBRE'], "RECURSOS HUMANOS", 'RECIBIO_NOMBRE')
        
        # =============================================================
        # GUARDAR Y CERRAR
        # =============================================================
        
        print(f"[MOVPER EXCEL] Guardando archivo...")
        wb.Save()
        
        print(f"[MOVPER EXCEL] Cerrando workbook...")
        wb.Close(SaveChanges=False)  # Ya guardamos con Save()
        
        print(f"[MOVPER EXCEL] Archivo guardado exitosamente: {output_path}")
        print(f"[MOVPER EXCEL] Verificando que el archivo existe...")
        if os.path.exists(output_path):
            print(f"[MOVPER EXCEL] [OK] Archivo confirmado: {os.path.getsize(output_path)} bytes")
        else:
            print(f"[MOVPER EXCEL] [ERROR] Archivo no existe después de guardar")
        
    except Exception as e:
        print(f"[MOVPER EXCEL] ERROR durante generacion: {e}")
        import traceback
        import sys
        traceback.print_exc(file=sys.stderr)
        try:
            wb.Close(SaveChanges=False)
        except:
            pass
        raise
    finally:
        print(f"[MOVPER EXCEL] Cerrando Excel...")
        excel.Quit()
        pythoncom.CoUninitialize()
        print(f"[MOVPER EXCEL] Excel cerrado")
    
    return output_path, output_filename


def construir_fechas_aplicacion(incidencias: List[Dict], quincena: Dict) -> str:
    """
    Construye la cadena de fechas de aplicación.
    Formato compacto: sin espacios si hay más de 4 días.
    Ejemplo con pocos días: "3, 5, 8 de febrero"
    Ejemplo con muchos días: "1,2,3,5,6,7,8,9,12,15 de febrero"
    """
    if not incidencias:
        return ""
    
    # Extraer días únicos
    dias = sorted(set(inc['fecha'].day for inc in incidencias if inc.get('fecha')))
    mes = quincena['inicio'].month
    mes_nombre = MESES_ES.get(mes, '')
    
    # Formato compacto: sin espacios si hay más de 4 días
    if len(dias) <= 4:
        dias_str = ', '.join(str(d) for d in dias)
    else:
        dias_str = ','.join(str(d) for d in dias)
    
    return f"{dias_str} de {mes_nombre}"


def analizar_tipos_incidencias(incidencias: List[Dict]) -> Dict:
    """
    Analiza las incidencias y agrupa por tipo.
    Usa estado_auto para retardos y clasificacion para faltas.
    Las faltas se agrupan por su clasificación específica (REMOTO, GUARDIA, etc.)
    También detecta salidas tempranas y olvido de checada.
    """
    tipos = {
        'faltas': [],           # Faltas sin clasificar o clasificadas como FALTA genérica
        'retardos': [],
        'remotos': [],          # Faltas clasificadas como TRABAJO_REMOTO/REMOTO
        'guardias': [],         # Faltas clasificadas como GUARDIA/GUARDIA_TELEFONICA
        'permisos': [],         # Faltas clasificadas como PERMISO
        'vacaciones': [],       # Faltas clasificadas como VACACIONES
        'incapacidades': [],    # Faltas clasificadas como INCAPACIDAD (no se usa en motivo)
        'olvido_checar': [],
        'salir_regresar': [],
        'retirarse_temprano': [],
    }
    
    for inc in incidencias:
        estado_auto = inc.get('estado_auto', '')
        clasificacion = inc.get('clasificacion', '')
        fecha = inc.get('fecha')
        dia = fecha.day if fecha else 0
        
        if estado_auto == 'RETARDO':
            tipos['retardos'].append(dia)
        elif estado_auto == 'FALTA':
            # Agrupar faltas por su clasificación específica
            if clasificacion in ('REMOTO', 'TRABAJO_REMOTO'):
                tipos['remotos'].append(dia)
            elif clasificacion in ('GUARDIA', 'GUARDIA_TELEFONICA'):
                tipos['guardias'].append(dia)
            elif clasificacion == 'PERMISO':
                tipos['permisos'].append(dia)
            elif clasificacion == 'VACACIONES':
                tipos['vacaciones'].append(dia)
            elif clasificacion == 'INCAPACIDAD':
                tipos['incapacidades'].append(dia)
            elif clasificacion == 'ERROR_SISTEMA':
                pass  # No es falta real, no incluir en ninguna categoría
            else:
                tipos['faltas'].append(dia)
        
        # Salida temprana justificada
        salida_estado = inc.get('salida_estado', '')
        salida_justificado = inc.get('salida_justificado', True)
        if salida_estado == 'SALIDA_TEMPRANA' and salida_justificado:
            tipos['retirarse_temprano'].append(dia)
        
        # Olvido de checada (entrada o salida) — solo si justificado
        entrada_nc = inc.get('entrada_no_checada', False)
        salida_nc = inc.get('salida_no_checada', False)
        olvido_just = inc.get('olvido_checar_justificado', True)
        if (entrada_nc or salida_nc) and olvido_just:
            tipos['olvido_checar'].append(dia)
    
    return tipos


def generar_texto_motivo(incidencias: List[Dict], tipos: Dict) -> str:
    """
    Genera el texto del motivo agrupando por tipo.
    Cada categoría es independiente y no se mezcla.
    
    Ejemplo: "1,2,3,4,5,6 retardo justificado. 7 falta justificada, trabajo remoto. 
              8 falta justificada, guardia telefónico."
    """
    partes = []
    
    # Retardos justificados
    if tipos['retardos']:
        dias_str = agrupar_dias_consecutivos(tipos['retardos'])
        partes.append(f"{dias_str} retardo justificado")
    
    # Faltas genéricas (sin clasificación específica)
    if tipos['faltas']:
        dias_str = agrupar_dias_consecutivos(tipos['faltas'])
        partes.append(f"{dias_str} falta justificada")
    
    # Faltas por trabajo remoto
    if tipos['remotos']:
        dias_str = agrupar_dias_consecutivos(tipos['remotos'])
        partes.append(f"{dias_str} falta justificada, trabajo remoto")
    
    # Faltas por guardia telefónica
    if tipos['guardias']:
        dias_str = agrupar_dias_consecutivos(tipos['guardias'])
        partes.append(f"{dias_str} falta justificada, guardia telefónico")
    
    # Faltas por permiso
    if tipos['permisos']:
        dias_str = agrupar_dias_consecutivos(tipos['permisos'])
        partes.append(f"{dias_str} falta justificada, permiso")
    
    # Faltas por vacaciones
    if tipos['vacaciones']:
        dias_str = agrupar_dias_consecutivos(tipos['vacaciones'])
        partes.append(f"{dias_str} falta justificada, vacaciones")
    
    # Salidas tempranas justificadas
    if tipos['retirarse_temprano']:
        dias_str = agrupar_dias_consecutivos(tipos['retirarse_temprano'])
        partes.append(f"{dias_str} salida temprana justificada")
    
    # Olvido de checada (auto-justificado)
    if tipos['olvido_checar']:
        dias_str = agrupar_dias_consecutivos(tipos['olvido_checar'])
        partes.append(f"{dias_str} olvido checar justificado")
    
    # NOTA: Incapacidades NO se incluyen en el motivo del MobPer
    
    return ". ".join(partes) + "." if partes else ""


def marcar_circulo_activo(sheet, nombre_celda: str):
    """
    Marca un círculo como activo en el Excel.
    Los círculos en el template son shapes; por ahora marcamos con ●
    """
    if nombre_celda in CELDAS:
        # Nota: En el template real, aquí se manipularía el shape
        # Por ahora dejamos sin cambios ya que los shapes ya existen
        pass


def excel_to_html_preview(excel_path: str) -> str:
    """
    Convierte un archivo Excel a HTML para vista previa en navegador.
    Usa openpyxl para leer los datos.
    
    Returns:
        str: Contenido HTML del Excel
    """
    return generar_html_preview_fallback(excel_path, "")


def generar_html_preview_fallback(excel_path: str, error_msg: str = "") -> str:
    """
    Genera un preview HTML básico.
    Usa openpyxl para leer los datos.
    """
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(excel_path)
        ws = wb.active
        
        html = ['<style>']
        html.append('table { border-collapse: collapse; width: 100%; font-family: Arial, sans-serif; }')
        html.append('td, th { border: 1px solid #ddd; padding: 8px; text-align: left; }')
        html.append('th { background: #5D4037; color: white; }')
        html.append('tr:nth-child(even) { background: #f9f9f9; }')
        html.append('.merged { font-weight: bold; }')
        html.append('</style>')
        html.append('<table>')
        
        for row in ws.iter_rows(min_row=1, max_row=40, max_col=10):
            html.append('<tr>')
            for cell in row:
                value = cell.value if cell.value else ''
                html.append(f'<td>{value}</td>')
            html.append('</tr>')
        
        html.append('</table>')
        
        if error_msg:
            html.append(f'<p style="color: #888; font-size: 12px; margin-top: 20px;">Nota: Vista simplificada. {error_msg}</p>')
        
        return '\n'.join(html)
        
    except Exception as e:
        return f'<div style="padding: 20px; background: #fff3cd; border: 1px solid #ffc107; border-radius: 8px;"><strong>Error al generar vista previa:</strong><br>{error_msg}<br><br>Error adicional: {str(e)}</div>'


def obtener_formatos_generados(user_id: int) -> List[Dict]:
    """
    Lista los formatos Excel generados para un usuario.
    """
    archivos = []
    
    if os.path.exists(OUTPUT_DIR):
        for filename in os.listdir(OUTPUT_DIR):
            if filename.endswith('.xlsx'):
                filepath = os.path.join(OUTPUT_DIR, filename)
                stat = os.stat(filepath)
                archivos.append({
                    'filename': filename,
                    'path': filepath,
                    'size': stat.st_size,
                    'created': datetime.fromtimestamp(stat.st_mtime)
                })
    
    # Ordenar por fecha de creación (más reciente primero)
    archivos.sort(key=lambda x: x['created'], reverse=True)
    
    return archivos
